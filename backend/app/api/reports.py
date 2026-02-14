image.png"""
报表API
- 财务报表
- 本月报表
- 本季度报表
- 本年度报表
"""
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, extract
from typing import Optional
from datetime import date, datetime, timedelta
from decimal import Decimal
import io
import calendar

from app.database import get_db
from app.models.user import User
from app.models.affiliate_account import AffiliateAccount, AffiliatePlatform
from app.models.affiliate_transaction import AffiliateTransaction
from app.models.google_ads_api_data import GoogleAdsApiData, GoogleMccAccount
from app.middleware.auth import get_current_user, get_current_manager_or_leader

router = APIRouter(prefix="/api/reports", tags=["reports"])

# 平台代码映射（支持新旧两种格式：小写缩写和全称）
PLATFORM_SHORT = {
    # 新格式（小写缩写）
    'cg': 'CG',
    'rw': 'RW',
    'lh': 'LH',
    'pb': 'PB',
    'lb': 'LB',
    'pm': 'PM',
    'bsh': 'BSH',
    'cf': 'CF',
    # 旧格式（全称，兼容历史数据）
    'collabglow': 'CG',
    'rewardoo': 'RW',
    'linkhaitao': 'LH',
    'partnerboost': 'PB',
    'linkbux': 'LB',
    'partnermatic': 'PM',
    'brandsparkhub': 'BSH',
    'creatorflare': 'CF',
}

# wj07的RW账号合并映射（3个账号合并为wenjun）
WJ07_RW_MERGE = ['wenjun', 'thegoodsandguard', 'vivaluxelife']


def get_platform_short_code(platform_code: str) -> str:
    """获取平台简短代码"""
    if not platform_code:
        return "未知"
    for key, short in PLATFORM_SHORT.items():
        if key in platform_code.lower():
            return short
    return platform_code


def get_date_range(period: str, year: int = None, month: int = None, quarter: int = None):
    """根据周期获取日期范围"""
    today = date.today()
    
    if period == "month":
        # 本月
        if year and month:
            start = date(year, month, 1)
            if month == 12:
                end = date(year + 1, 1, 1)
            else:
                end = date(year, month + 1, 1)
            end = date(end.year, end.month, 1) - timedelta(days=1) if end.day == 1 else end
            # 简化：直接用下月1日减1天
            import calendar
            last_day = calendar.monthrange(year, month)[1]
            end = date(year, month, last_day)
        else:
            start = date(today.year, today.month, 1)
            import calendar
            last_day = calendar.monthrange(today.year, today.month)[1]
            end = date(today.year, today.month, last_day)
    elif period == "quarter":
        # 本季度
        if year and quarter:
            start_month = (quarter - 1) * 3 + 1
            end_month = quarter * 3
            start = date(year, start_month, 1)
            import calendar
            last_day = calendar.monthrange(year, end_month)[1]
            end = date(year, end_month, last_day)
        else:
            current_quarter = (today.month - 1) // 3 + 1
            start_month = (current_quarter - 1) * 3 + 1
            end_month = current_quarter * 3
            start = date(today.year, start_month, 1)
            import calendar
            last_day = calendar.monthrange(today.year, end_month)[1]
            end = date(today.year, end_month, last_day)
    elif period == "year":
        # 本年度
        if year:
            start = date(year, 1, 1)
            end = date(year, 12, 31)
        else:
            start = date(today.year, 1, 1)
            end = date(today.year, 12, 31)
    else:
        # 默认本月
        start = date(today.year, today.month, 1)
        import calendar
        last_day = calendar.monthrange(today.year, today.month)[1]
        end = date(today.year, today.month, last_day)
    
    return start, end


@router.get("/financial")
async def get_financial_report(
    year: int = Query(..., description="年份"),
    month: int = Query(..., description="月份"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    财务报表
    按员工 -> 平台 -> 账号 展示
    """
    import calendar
    start_date = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    end_date = date(year, month, last_day)
    
    # 获取所有员工
    employees = db.query(User).filter(User.role == 'employee').order_by(User.username).all()
    
    result = []
    total_ad_cost = Decimal('0')
    total_book_commission = Decimal('0')
    total_rejected_commission = Decimal('0')
    
    for emp in employees:
        emp_display_name = emp.display_name or emp.username
        
        # 1. 获取该员工的广告费（从google_ads_api_data）
        emp_ad_cost = db.query(func.sum(GoogleAdsApiData.cost)).filter(
            GoogleAdsApiData.user_id == emp.id,
            GoogleAdsApiData.date >= start_date,
            GoogleAdsApiData.date <= end_date
        ).scalar() or Decimal('0')
        
        # 货币转换（如果有CNY的MCC）
        mcc_accounts = db.query(GoogleMccAccount).filter(GoogleMccAccount.user_id == emp.id).all()
        cny_mcc_ids = [m.id for m in mcc_accounts if m.currency == 'CNY']
        
        if cny_mcc_ids:
            # 单独计算CNY的费用并转换
            cny_cost = db.query(func.sum(GoogleAdsApiData.cost)).filter(
                GoogleAdsApiData.user_id == emp.id,
                GoogleAdsApiData.mcc_id.in_(cny_mcc_ids),
                GoogleAdsApiData.date >= start_date,
                GoogleAdsApiData.date <= end_date
            ).scalar() or Decimal('0')
            
            usd_cost = db.query(func.sum(GoogleAdsApiData.cost)).filter(
                GoogleAdsApiData.user_id == emp.id,
                GoogleAdsApiData.mcc_id.notin_(cny_mcc_ids),
                GoogleAdsApiData.date >= start_date,
                GoogleAdsApiData.date <= end_date
            ).scalar() or Decimal('0')
            
            emp_ad_cost = Decimal(str(usd_cost)) + Decimal(str(cny_cost)) / Decimal('7.2')
        
        emp_ad_cost = Decimal(str(emp_ad_cost))
        total_ad_cost += emp_ad_cost
        
        # 2. 获取该员工的平台账号数据
        accounts = db.query(AffiliateAccount).join(AffiliatePlatform).filter(
            AffiliateAccount.user_id == emp.id,
            AffiliateAccount.is_active == True
        ).all()
        
        # 按平台分组
        platform_data = {}
        for acc in accounts:
            platform_code = get_platform_short_code(acc.platform.platform_code) if acc.platform else "未知"
            
            # wj07的RW账号合并处理
            account_name = acc.account_name
            if emp.username == 'wj07' and platform_code == 'RW':
                # 合并为wenjun
                account_name = 'wenjun'
                if platform_code in platform_data and account_name in [a['account_name'] for a in platform_data[platform_code]]:
                    # 已经添加过了，跳过但要累加数据
                    continue
            
            # 账面佣金（所有状态）
            book_commission = db.query(func.sum(AffiliateTransaction.commission_amount)).filter(
                AffiliateTransaction.affiliate_account_id == acc.id,
                AffiliateTransaction.transaction_time >= datetime.combine(start_date, datetime.min.time()),
                AffiliateTransaction.transaction_time <= datetime.combine(end_date, datetime.max.time())
            ).scalar() or Decimal('0')
            
            # 失效佣金（rejected状态）
            rejected_commission = db.query(func.sum(AffiliateTransaction.commission_amount)).filter(
                AffiliateTransaction.affiliate_account_id == acc.id,
                AffiliateTransaction.status == 'rejected',
                AffiliateTransaction.transaction_time >= datetime.combine(start_date, datetime.min.time()),
                AffiliateTransaction.transaction_time <= datetime.combine(end_date, datetime.max.time())
            ).scalar() or Decimal('0')
            
            book_commission = Decimal(str(book_commission))
            rejected_commission = Decimal(str(rejected_commission))
            
            if platform_code not in platform_data:
                platform_data[platform_code] = []
            
            # 检查是否需要合并（wj07的RW账号）
            if emp.username == 'wj07' and platform_code == 'RW':
                # 查找已存在的wenjun记录
                existing = None
                for item in platform_data[platform_code]:
                    if item['account_name'] == 'wenjun':
                        existing = item
                        break
                
                if existing:
                    existing['book_commission'] = Decimal(str(existing['book_commission'])) + book_commission
                    existing['rejected_commission'] = Decimal(str(existing['rejected_commission'])) + rejected_commission
                else:
                    platform_data[platform_code].append({
                        'account_name': 'wenjun',
                        'book_commission': float(book_commission),
                        'rejected_commission': float(rejected_commission)
                    })
            else:
                platform_data[platform_code].append({
                    'account_name': acc.account_name,
                    'book_commission': float(book_commission),
                    'rejected_commission': float(rejected_commission)
                })
        
        # 如果是wj07，需要额外处理：合并3个RW账号的佣金
        if emp.username == 'wj07' and 'RW' in platform_data:
            # 获取所有3个RW账号的ID
            rw_account_ids = [acc.id for acc in accounts if get_platform_short_code(acc.platform.platform_code) == 'RW']
            
            if rw_account_ids:
                # 重新计算合并后的佣金
                total_book = db.query(func.sum(AffiliateTransaction.commission_amount)).filter(
                    AffiliateTransaction.affiliate_account_id.in_(rw_account_ids),
                    AffiliateTransaction.transaction_time >= datetime.combine(start_date, datetime.min.time()),
                    AffiliateTransaction.transaction_time <= datetime.combine(end_date, datetime.max.time())
                ).scalar() or Decimal('0')
                
                total_rejected = db.query(func.sum(AffiliateTransaction.commission_amount)).filter(
                    AffiliateTransaction.affiliate_account_id.in_(rw_account_ids),
                    AffiliateTransaction.status == 'rejected',
                    AffiliateTransaction.transaction_time >= datetime.combine(start_date, datetime.min.time()),
                    AffiliateTransaction.transaction_time <= datetime.combine(end_date, datetime.max.time())
                ).scalar() or Decimal('0')
                
                platform_data['RW'] = [{
                    'account_name': 'wenjun',
                    'book_commission': float(total_book),
                    'rejected_commission': float(total_rejected)
                }]
        
        # 构建员工数据
        emp_accounts = []
        emp_book_commission = Decimal('0')
        emp_rejected_commission = Decimal('0')
        
        for platform_code, accounts_list in platform_data.items():
            for acc_data in accounts_list:
                emp_accounts.append({
                    'platform': platform_code,
                    'account_name': acc_data['account_name'],
                    'book_commission': acc_data['book_commission'],
                    'rejected_commission': acc_data['rejected_commission']
                })
                emp_book_commission += Decimal(str(acc_data['book_commission']))
                emp_rejected_commission += Decimal(str(acc_data['rejected_commission']))
        
        total_book_commission += emp_book_commission
        total_rejected_commission += emp_rejected_commission
        
        result.append({
            'employee': emp_display_name,
            'username': emp.username,
            'ad_cost': float(emp_ad_cost),
            'accounts': emp_accounts,
            'total_book_commission': float(emp_book_commission),
            'total_rejected_commission': float(emp_rejected_commission)
        })
    
    return {
        'year': year,
        'month': month,
        'data': result,
        'summary': {
            'total_ad_cost': float(total_ad_cost),
            'total_book_commission': float(total_book_commission),
            'total_rejected_commission': float(total_rejected_commission),
            'total_valid_commission': float(total_book_commission - total_rejected_commission)
        }
    }


@router.get("/monthly")
async def get_monthly_report(
    year: int = Query(None, description="年份"),
    month: int = Query(None, description="月份"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    本月报表
    按员工汇总
    """
    import calendar
    today = date.today()
    year = year or today.year
    month = month or today.month
    
    start_date = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    end_date = date(year, month, last_day)
    
    return await _get_summary_report(db, start_date, end_date, f"{year}年{month}月")


@router.get("/quarterly")
async def get_quarterly_report(
    year: int = Query(None, description="年份"),
    quarter: int = Query(None, description="季度 (1-4)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    本季度报表
    按员工汇总
    """
    import calendar
    today = date.today()
    year = year or today.year
    quarter = quarter or ((today.month - 1) // 3 + 1)
    
    start_month = (quarter - 1) * 3 + 1
    end_month = quarter * 3
    
    start_date = date(year, start_month, 1)
    last_day = calendar.monthrange(year, end_month)[1]
    end_date = date(year, end_month, last_day)
    
    return await _get_summary_report(db, start_date, end_date, f"{year}年Q{quarter}")


@router.get("/yearly")
async def get_yearly_report(
    year: int = Query(None, description="年份"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    本年度报表
    按员工汇总
    """
    today = date.today()
    year = year or today.year
    
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)
    
    return await _get_summary_report(db, start_date, end_date, f"{year}年")


async def _get_summary_report(db: Session, start_date: date, end_date: date, period_name: str):
    """
    生成汇总报表
    """
    employees = db.query(User).filter(User.role == 'employee').order_by(User.username).all()
    
    result = []
    total_ad_cost = Decimal('0')
    total_book_commission = Decimal('0')
    total_rejected_commission = Decimal('0')
    total_orders = 0
    total_active_campaigns = 0
    
    for emp in employees:
        emp_display_name = emp.display_name or emp.username
        
        # 1. 广告费
        emp_ad_cost = db.query(func.sum(GoogleAdsApiData.cost)).filter(
            GoogleAdsApiData.user_id == emp.id,
            GoogleAdsApiData.date >= start_date,
            GoogleAdsApiData.date <= end_date
        ).scalar() or Decimal('0')
        
        # 货币转换
        mcc_accounts = db.query(GoogleMccAccount).filter(GoogleMccAccount.user_id == emp.id).all()
        cny_mcc_ids = [m.id for m in mcc_accounts if m.currency == 'CNY']
        
        if cny_mcc_ids:
            cny_cost = db.query(func.sum(GoogleAdsApiData.cost)).filter(
                GoogleAdsApiData.user_id == emp.id,
                GoogleAdsApiData.mcc_id.in_(cny_mcc_ids),
                GoogleAdsApiData.date >= start_date,
                GoogleAdsApiData.date <= end_date
            ).scalar() or Decimal('0')
            
            usd_cost = db.query(func.sum(GoogleAdsApiData.cost)).filter(
                GoogleAdsApiData.user_id == emp.id,
                GoogleAdsApiData.mcc_id.notin_(cny_mcc_ids),
                GoogleAdsApiData.date >= start_date,
                GoogleAdsApiData.date <= end_date
            ).scalar() or Decimal('0')
            
            emp_ad_cost = Decimal(str(usd_cost)) + Decimal(str(cny_cost)) / Decimal('7.2')
        
        emp_ad_cost = Decimal(str(emp_ad_cost))
        total_ad_cost += emp_ad_cost
        
        # 2. 账面佣金（所有状态）
        emp_book_commission = db.query(func.sum(AffiliateTransaction.commission_amount)).filter(
            AffiliateTransaction.user_id == emp.id,
            AffiliateTransaction.transaction_time >= datetime.combine(start_date, datetime.min.time()),
            AffiliateTransaction.transaction_time <= datetime.combine(end_date, datetime.max.time())
        ).scalar() or Decimal('0')
        emp_book_commission = Decimal(str(emp_book_commission))
        total_book_commission += emp_book_commission
        
        # 3. 失效佣金（rejected状态）
        emp_rejected_commission = db.query(func.sum(AffiliateTransaction.commission_amount)).filter(
            AffiliateTransaction.user_id == emp.id,
            AffiliateTransaction.status == 'rejected',
            AffiliateTransaction.transaction_time >= datetime.combine(start_date, datetime.min.time()),
            AffiliateTransaction.transaction_time <= datetime.combine(end_date, datetime.max.time())
        ).scalar() or Decimal('0')
        emp_rejected_commission = Decimal(str(emp_rejected_commission))
        total_rejected_commission += emp_rejected_commission
        
        # 4. 订单数
        emp_orders = db.query(func.count(AffiliateTransaction.id)).filter(
            AffiliateTransaction.user_id == emp.id,
            AffiliateTransaction.transaction_time >= datetime.combine(start_date, datetime.min.time()),
            AffiliateTransaction.transaction_time <= datetime.combine(end_date, datetime.max.time())
        ).scalar() or 0
        total_orders += emp_orders
        
        # 5. 在跑广告量（已启用的广告系列数）
        # 取最新一天的数据
        emp_active_campaigns = db.query(func.count(func.distinct(GoogleAdsApiData.campaign_id))).filter(
            GoogleAdsApiData.user_id == emp.id,
            GoogleAdsApiData.status == '已启用',
            GoogleAdsApiData.date == end_date
        ).scalar() or 0
        
        # 如果最后一天没有数据，往前找
        if emp_active_campaigns == 0:
            latest_date = db.query(func.max(GoogleAdsApiData.date)).filter(
                GoogleAdsApiData.user_id == emp.id,
                GoogleAdsApiData.date <= end_date
            ).scalar()
            
            if latest_date:
                emp_active_campaigns = db.query(func.count(func.distinct(GoogleAdsApiData.campaign_id))).filter(
                    GoogleAdsApiData.user_id == emp.id,
                    GoogleAdsApiData.status == '已启用',
                    GoogleAdsApiData.date == latest_date
                ).scalar() or 0
        
        total_active_campaigns += emp_active_campaigns
        
        # 有效佣金
        emp_valid_commission = emp_book_commission - emp_rejected_commission
        
        result.append({
            'employee': emp_display_name,
            'username': emp.username,
            'ad_cost': float(emp_ad_cost),
            'book_commission': float(emp_book_commission),
            'rejected_commission': float(emp_rejected_commission),
            'valid_commission': float(emp_valid_commission),
            'orders': emp_orders,
            'active_campaigns': emp_active_campaigns
        })
    
    return {
        'period': period_name,
        'start_date': start_date.isoformat(),
        'end_date': end_date.isoformat(),
        'data': result,
        'summary': {
            'total_ad_cost': float(total_ad_cost),
            'total_book_commission': float(total_book_commission),
            'total_rejected_commission': float(total_rejected_commission),
            'total_valid_commission': float(total_book_commission - total_rejected_commission),
            'total_orders': total_orders,
            'total_active_campaigns': total_active_campaigns
        }
    }


@router.get("/monthly/export")
async def export_monthly_report(
    year: int = Query(None, description="年份"),
    month: int = Query(None, description="月份"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    导出月度报表为Excel
    """
    today = date.today()
    year = year or today.year
    month = month or today.month
    
    start_date = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    end_date = date(year, month, last_day)
    
    report_data = await _get_summary_report(db, start_date, end_date, f"{year}年{month}月")
    
    return _generate_summary_excel(report_data, f"月度报表_{year}年{month}月.xlsx")


@router.get("/quarterly/export")
async def export_quarterly_report(
    year: int = Query(None, description="年份"),
    quarter: int = Query(None, description="季度 (1-4)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    导出季度报表为Excel
    """
    today = date.today()
    year = year or today.year
    quarter = quarter or ((today.month - 1) // 3 + 1)
    
    start_month = (quarter - 1) * 3 + 1
    end_month = quarter * 3
    
    start_date = date(year, start_month, 1)
    last_day = calendar.monthrange(year, end_month)[1]
    end_date = date(year, end_month, last_day)
    
    report_data = await _get_summary_report(db, start_date, end_date, f"{year}年Q{quarter}")
    
    return _generate_summary_excel(report_data, f"季度报表_{year}年Q{quarter}.xlsx")


@router.get("/yearly/export")
async def export_yearly_report(
    year: int = Query(None, description="年份"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    导出年度报表为Excel
    """
    today = date.today()
    year = year or today.year
    
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)
    
    report_data = await _get_summary_report(db, start_date, end_date, f"{year}年")
    
    return _generate_summary_excel(report_data, f"年度报表_{year}年.xlsx")


def _generate_summary_excel(report_data: dict, filename: str) -> StreamingResponse:
    """
    生成汇总报表Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from urllib.parse import quote
    
    wb = Workbook()
    ws = wb.active
    ws.title = "报表"
    
    # 样式
    header_fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
    header_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题行
    ws.merge_cells('A1:G1')
    ws['A1'] = f"{report_data.get('period', '')} 报表"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_align
    
    # 日期范围
    ws.merge_cells('A2:G2')
    ws['A2'] = f"日期范围: {report_data.get('start_date', '')} ~ {report_data.get('end_date', '')}"
    ws['A2'].alignment = center_align
    
    # 表头
    headers = ['员工', '广告费($)', '账面佣金($)', '失效佣金($)', '有效佣金($)', '订单数', '在跑广告量']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # 数据行
    data = report_data.get('data', [])
    for row_idx, row_data in enumerate(data, 5):
        ws.cell(row=row_idx, column=1, value=row_data.get('employee', '')).border = thin_border
        ws.cell(row=row_idx, column=2, value=round(row_data.get('ad_cost', 0), 2)).border = thin_border
        ws.cell(row=row_idx, column=3, value=round(row_data.get('book_commission', 0), 2)).border = thin_border
        ws.cell(row=row_idx, column=4, value=round(row_data.get('rejected_commission', 0), 2)).border = thin_border
        ws.cell(row=row_idx, column=5, value=round(row_data.get('valid_commission', 0), 2)).border = thin_border
        ws.cell(row=row_idx, column=6, value=row_data.get('orders', 0)).border = thin_border
        ws.cell(row=row_idx, column=7, value=row_data.get('active_campaigns', 0)).border = thin_border
        
        # 居中数字
        for col in range(2, 8):
            ws.cell(row=row_idx, column=col).alignment = center_align
    
    # 合计行
    summary = report_data.get('summary', {})
    summary_row = 5 + len(data)
    summary_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    
    ws.cell(row=summary_row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=summary_row, column=1).fill = summary_fill
    ws.cell(row=summary_row, column=1).border = thin_border
    
    ws.cell(row=summary_row, column=2, value=round(summary.get('total_ad_cost', 0), 2)).fill = summary_fill
    ws.cell(row=summary_row, column=2).border = thin_border
    ws.cell(row=summary_row, column=2).font = Font(bold=True)
    
    ws.cell(row=summary_row, column=3, value=round(summary.get('total_book_commission', 0), 2)).fill = summary_fill
    ws.cell(row=summary_row, column=3).border = thin_border
    ws.cell(row=summary_row, column=3).font = Font(bold=True)
    
    ws.cell(row=summary_row, column=4, value=round(summary.get('total_rejected_commission', 0), 2)).fill = summary_fill
    ws.cell(row=summary_row, column=4).border = thin_border
    ws.cell(row=summary_row, column=4).font = Font(bold=True)
    
    ws.cell(row=summary_row, column=5, value=round(summary.get('total_valid_commission', 0), 2)).fill = summary_fill
    ws.cell(row=summary_row, column=5).border = thin_border
    ws.cell(row=summary_row, column=5).font = Font(bold=True)
    
    ws.cell(row=summary_row, column=6, value=summary.get('total_orders', 0)).fill = summary_fill
    ws.cell(row=summary_row, column=6).border = thin_border
    ws.cell(row=summary_row, column=6).font = Font(bold=True)
    
    ws.cell(row=summary_row, column=7, value=summary.get('total_active_campaigns', 0)).fill = summary_fill
    ws.cell(row=summary_row, column=7).border = thin_border
    ws.cell(row=summary_row, column=7).font = Font(bold=True)
    
    for col in range(2, 8):
        ws.cell(row=summary_row, column=col).alignment = center_align
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )


@router.get("/financial/export")
async def export_financial_report(
    year: int = Query(..., description="年份"),
    month: int = Query(..., description="月份"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    导出财务报表为Excel格式
    格式与2026年丰度收支统计表一致
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from urllib.parse import quote
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="服务器未安装openpyxl库")
    
    # 获取数据
    start_date = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    end_date = date(year, month, last_day)
    
    # 获取所有员工
    employees = db.query(User).filter(User.role == 'employee').order_by(User.username).all()
    
    # 平台列表
    PLATFORMS = ['CG', 'RW', 'LH', 'PM', 'LB', 'PB', 'BSH', 'CF']
    NUM_PLATFORMS = len(PLATFORMS)
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month:02d}月"
    
    # 样式定义
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    bold_font = Font(bold=True)
    red_font = Font(color="FF0000")
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 行标题定义（不包含在跑广告量，它在列中）
    ROW_LABELS = [
        '月份',           # 1
        'MCC',            # 2  
        '币种',           # 3
        '广告费',         # 4
        '广告联盟',       # 5
        '账号名称',       # 6
        '账面佣金（美金）',  # 7
        '失效佣金（美金）',  # 8
        '应收佣金（美金）-5号',   # 9
        '应收佣金（美金）-15号',  # 10
        '应收佣金（美金）-合计',  # 11
        '实收佣金（人民币）-10号', # 12
        '实收佣金（人民币）-20号', # 13
        '实收佣金（人民币）-合计', # 14
        '收款人',          # 15
        '收款卡号',        # 16
        '可分配利润（实收佣金-广告费）', # 17
    ]
    
    # 写入A列标题
    for i, label in enumerate(ROW_LABELS, start=1):
        cell = ws.cell(row=i, column=1, value=label)
        cell.border = thin_border
        if i <= 2:
            cell.fill = yellow_fill
            cell.font = bold_font
    
    # 收集所有员工数据
    all_emp_data = []
    total_ad_cost = Decimal('0')
    total_active_campaigns = 0
    platform_totals_book = {p: Decimal('0') for p in PLATFORMS}
    platform_totals_rejected = {p: Decimal('0') for p in PLATFORMS}
    
    for emp in employees:
        emp_display_name = emp.display_name or emp.username
        
        # 广告费
        emp_ad_cost = db.query(func.sum(GoogleAdsApiData.cost)).filter(
            GoogleAdsApiData.user_id == emp.id,
            GoogleAdsApiData.date >= start_date,
            GoogleAdsApiData.date <= end_date
        ).scalar() or Decimal('0')
        
        # 货币转换
        mcc_accounts = db.query(GoogleMccAccount).filter(GoogleMccAccount.user_id == emp.id).all()
        cny_mcc_ids = [m.id for m in mcc_accounts if m.currency == 'CNY']
        
        if cny_mcc_ids:
            cny_cost = db.query(func.sum(GoogleAdsApiData.cost)).filter(
                GoogleAdsApiData.user_id == emp.id,
                GoogleAdsApiData.mcc_id.in_(cny_mcc_ids),
                GoogleAdsApiData.date >= start_date,
                GoogleAdsApiData.date <= end_date
            ).scalar() or Decimal('0')
            
            usd_cost = db.query(func.sum(GoogleAdsApiData.cost)).filter(
                GoogleAdsApiData.user_id == emp.id,
                GoogleAdsApiData.mcc_id.notin_(cny_mcc_ids),
                GoogleAdsApiData.date >= start_date,
                GoogleAdsApiData.date <= end_date
            ).scalar() or Decimal('0')
            
            emp_ad_cost = Decimal(str(usd_cost)) + Decimal(str(cny_cost)) / Decimal('7.2')
        
        emp_ad_cost = Decimal(str(emp_ad_cost))
        total_ad_cost += emp_ad_cost
        
        # 在跑广告量
        active_campaigns = db.query(func.count(func.distinct(GoogleAdsApiData.campaign_id))).filter(
            GoogleAdsApiData.user_id == emp.id,
            GoogleAdsApiData.status == '已启用',
            GoogleAdsApiData.date == end_date
        ).scalar() or 0
        
        if active_campaigns == 0:
            latest_date = db.query(func.max(GoogleAdsApiData.date)).filter(
                GoogleAdsApiData.user_id == emp.id,
                GoogleAdsApiData.date <= end_date
            ).scalar()
            if latest_date:
                active_campaigns = db.query(func.count(func.distinct(GoogleAdsApiData.campaign_id))).filter(
                    GoogleAdsApiData.user_id == emp.id,
                    GoogleAdsApiData.status == '已启用',
                    GoogleAdsApiData.date == latest_date
                ).scalar() or 0
        
        total_active_campaigns += active_campaigns
        
        # 平台账号数据
        accounts = db.query(AffiliateAccount).join(AffiliatePlatform).filter(
            AffiliateAccount.user_id == emp.id,
            AffiliateAccount.is_active == True
        ).all()
        
        platform_data = {p: {'names': [], 'book': Decimal('0'), 'rejected': Decimal('0'), 'payee_name': '', 'payee_card': ''} for p in PLATFORMS}
        
        for acc in accounts:
            platform_code = get_platform_short_code(acc.platform.platform_code) if acc.platform else None
            if platform_code not in PLATFORMS:
                continue
            
            book_commission = db.query(func.sum(AffiliateTransaction.commission_amount)).filter(
                AffiliateTransaction.affiliate_account_id == acc.id,
                AffiliateTransaction.transaction_time >= datetime.combine(start_date, datetime.min.time()),
                AffiliateTransaction.transaction_time <= datetime.combine(end_date, datetime.max.time())
            ).scalar() or Decimal('0')
            
            rejected_commission = db.query(func.sum(AffiliateTransaction.commission_amount)).filter(
                AffiliateTransaction.affiliate_account_id == acc.id,
                AffiliateTransaction.status == 'rejected',
                AffiliateTransaction.transaction_time >= datetime.combine(start_date, datetime.min.time()),
                AffiliateTransaction.transaction_time <= datetime.combine(end_date, datetime.max.time())
            ).scalar() or Decimal('0')
            
            platform_data[platform_code]['names'].append(acc.account_name)
            platform_data[platform_code]['book'] += Decimal(str(book_commission))
            platform_data[platform_code]['rejected'] += Decimal(str(rejected_commission))
            if acc.payee_name:
                platform_data[platform_code]['payee_name'] = acc.payee_name
            if acc.payee_card:
                platform_data[platform_code]['payee_card'] = acc.payee_card
            
            # 累加到总计
            platform_totals_book[platform_code] += Decimal(str(book_commission))
            platform_totals_rejected[platform_code] += Decimal(str(rejected_commission))
        
        all_emp_data.append({
            'name': emp_display_name,
            'ad_cost': emp_ad_cost,
            'active_campaigns': active_campaigns,
            'platform_data': platform_data
        })
    
    # === 写入数据 ===
    # 每个员工区块 = 7个平台 + 1个"在跑广告量"列 = 8列
    EMP_COLS = NUM_PLATFORMS + 1  # 8列
    current_col = 2
    
    # ========== 合计区块 ==========
    # 行2: 合计（合并7个平台列）
    ws.cell(row=2, column=current_col, value='合计')
    ws.merge_cells(start_row=2, start_column=current_col, end_row=2, end_column=current_col + NUM_PLATFORMS - 1)
    ws.cell(row=2, column=current_col).alignment = center_align
    ws.cell(row=2, column=current_col).fill = yellow_fill
    ws.cell(row=2, column=current_col).font = bold_font
    
    # 行2: 在跑广告量列标题
    active_col = current_col + NUM_PLATFORMS
    cell = ws.cell(row=2, column=active_col, value='在跑广告量')
    cell.fill = yellow_fill
    cell.font = bold_font
    cell.alignment = center_align
    
    # 行3: 美金（7列）+ 在跑广告量列空白
    for i in range(NUM_PLATFORMS):
        cell = ws.cell(row=3, column=current_col + i, value='美金')
        cell.fill = green_fill
        cell.alignment = center_align
    ws.cell(row=3, column=active_col, value='').fill = green_fill
    
    # 行4: 总广告费 + 总在跑广告量
    ws.cell(row=4, column=current_col, value=round(float(total_ad_cost), 2))
    ws.cell(row=4, column=active_col, value=total_active_campaigns)
    
    # 行5: 平台名称
    for i, platform in enumerate(PLATFORMS):
        cell = ws.cell(row=5, column=current_col + i, value=platform)
        cell.fill = green_fill
        cell.alignment = center_align
    ws.cell(row=5, column=active_col, value='').fill = green_fill
    
    # 行6: 账号名称（合计列为空）
    
    # 行7: 账面佣金汇总
    for i, platform in enumerate(PLATFORMS):
        ws.cell(row=7, column=current_col + i, value=round(float(platform_totals_book[platform]), 2))
    
    # 行8: 失效佣金汇总
    for i, platform in enumerate(PLATFORMS):
        ws.cell(row=8, column=current_col + i, value=round(float(platform_totals_rejected[platform]), 2))
    
    # 行9-14: 应收/实收佣金（预留占位，写0）
    for row in range(9, 15):
        for i in range(NUM_PLATFORMS):
            ws.cell(row=row, column=current_col + i, value=0)
    
    # 行17: 可分配利润
    profit = -float(total_ad_cost)  # 暂时用负广告费
    cell = ws.cell(row=17, column=current_col, value=round(profit, 2))
    if profit < 0:
        cell.font = red_font
    
    # 应用绿色背景到数据区（包括在跑广告量列）
    for row in range(3, 18):
        for i in range(EMP_COLS):
            cell = ws.cell(row=row, column=current_col + i)
            if cell.value is None:
                cell.value = ''
            cell.fill = green_fill
            cell.border = thin_border
    
    current_col += EMP_COLS
    
    # ========== 各员工数据 ==========
    for emp_data in all_emp_data:
        # 行2: 员工名（合并7个平台列）
        ws.cell(row=2, column=current_col, value=emp_data['name'])
        ws.merge_cells(start_row=2, start_column=current_col, end_row=2, end_column=current_col + NUM_PLATFORMS - 1)
        ws.cell(row=2, column=current_col).alignment = center_align
        ws.cell(row=2, column=current_col).fill = yellow_fill
        ws.cell(row=2, column=current_col).font = bold_font
        
        # 行2: 在跑广告量列标题
        active_col = current_col + NUM_PLATFORMS
        cell = ws.cell(row=2, column=active_col, value='在跑广告量')
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.alignment = center_align
        
        # 行3: 美金
        for i in range(NUM_PLATFORMS):
            cell = ws.cell(row=3, column=current_col + i, value='美金')
            cell.fill = green_fill
            cell.alignment = center_align
        ws.cell(row=3, column=active_col, value='').fill = green_fill
        
        # 行4: 广告费 + 在跑广告量
        ws.cell(row=4, column=current_col, value=round(float(emp_data['ad_cost']), 2))
        ws.cell(row=4, column=active_col, value=emp_data['active_campaigns'])
        
        # 行5: 平台名称
        for i, platform in enumerate(PLATFORMS):
            cell = ws.cell(row=5, column=current_col + i, value=platform)
            cell.fill = green_fill
            cell.alignment = center_align
        ws.cell(row=5, column=active_col, value='').fill = green_fill
        
        # 行6-8: 账号、佣金
        for i, platform in enumerate(PLATFORMS):
            pdata = emp_data['platform_data'][platform]
            col = current_col + i
            
            # 账号名
            ws.cell(row=6, column=col, value=','.join(pdata['names']) if pdata['names'] else '')
            # 账面佣金
            ws.cell(row=7, column=col, value=round(float(pdata['book']), 2))
            # 失效佣金
            ws.cell(row=8, column=col, value=round(float(pdata['rejected']), 2))
        
        # 行9-14: 预留占位
        for row in range(9, 15):
            for i in range(NUM_PLATFORMS):
                ws.cell(row=row, column=current_col + i, value=0)
        
        # 行15-16: 收款人/卡号（取第一个有值的平台）
        payee_name = ''
        payee_card = ''
        for platform in PLATFORMS:
            pdata = emp_data['platform_data'][platform]
            if pdata['payee_name'] and not payee_name:
                payee_name = pdata['payee_name']
            if pdata['payee_card'] and not payee_card:
                payee_card = pdata['payee_card']
        ws.cell(row=15, column=current_col, value=payee_name)
        ws.cell(row=16, column=current_col, value=payee_card)
        
        # 行17: 可分配利润
        emp_profit = -float(emp_data['ad_cost'])  # 暂时用负广告费
        cell = ws.cell(row=17, column=current_col, value=round(emp_profit, 2))
        if emp_profit < 0:
            cell.font = red_font
        
        # 应用绿色背景（包括在跑广告量列）
        for row in range(3, 18):
            for i in range(EMP_COLS):
                cell = ws.cell(row=row, column=current_col + i)
                if cell.value is None:
                    cell.value = ''
                cell.fill = green_fill
                cell.border = thin_border
        
        current_col += EMP_COLS
    
    # 设置列宽
    ws.column_dimensions['A'].width = 28
    for col in range(2, current_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 返回文件
    filename = f"财务报表_{year}年{month:02d}月.xlsx"
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )

