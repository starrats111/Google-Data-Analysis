"""
推荐商家 API — Excel 上传、查询、批次列表
"""
import uuid
import logging
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.database import get_db
from app.middleware.auth import get_current_user, get_current_manager_or_leader
from app.models.merchant import AffiliateMerchant
from app.models.merchant_recommendation import MerchantRecommendation
from app.models.user import User

router = APIRouter(prefix="/api/merchant-recommendations", tags=["推荐商家"])
logger = logging.getLogger(__name__)

PLATFORM_NAME_MAP = {
    "linkhaitao": "LH", "lh": "LH",
    "rewardoo": "RW", "rw": "RW",
    "collabglow": "CG", "cg": "CG",
    "brandsparkhub": "BSH", "bsh": "BSH",
    "partnermatic": "PM", "pm": "PM",
    "creatorflare": "CF", "cf": "CF",
    "linkbux": "LB", "lb": "LB",
}


def _clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _to_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


@router.post("/upload")
async def upload_recommendations(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_manager_or_leader),
    db: Session = Depends(get_db),
):
    """上传推荐商家 Excel，解析入库，标记推荐商家"""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")

    import openpyxl
    from io import BytesIO

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="文件大小不能超过 10MB")

    try:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法读取 Excel: {e}")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 无数据行")

    # 找到表头行（可能第一行是标题，第二行才是表头）
    header_row_idx = 0
    for i, row in enumerate(rows[:3]):
        cells = [str(c or "").strip().lower() for c in row]
        if "mcid" in cells or "mid" in cells:
            header_row_idx = i
            break

    headers = [str(h or "").strip() for h in rows[header_row_idx]]
    col = {}
    for i, h in enumerate(headers):
        hl = h.lower().replace("\n", " ")
        if hl == "mcid":
            col["mcid"] = i
        elif hl == "mid":
            col["merchant_mid"] = i
        elif "广告主名称" in h or "advertiser name" in hl:
            col["merchant_name"] = i
        elif "网址" in h or "website" in hl:
            col["merchant_url"] = i
        elif "商家地区" in h or "merchant base" in hl:
            col["merchant_region"] = i
        elif hl.startswith("epc"):
            col["epc"] = i
        elif "上限" in h or "cap" in hl:
            col["commission_cap"] = i
        elif "平均佣金比例" in h or "average commission rate" in hl:
            col["avg_commission_rate"] = i
        elif "平均客单佣金" in h or "average order commission" in hl:
            col["avg_order_commission"] = i
        elif "平台" in h or "platform" in hl:
            col["platform"] = i

    if "merchant_name" not in col:
        raise HTTPException(status_code=400, detail="Excel 缺少广告主名称列")

    batch_id = f"REC-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6]}"
    recommendations = []

    for row in rows[header_row_idx + 1:]:
        if not row or all(c is None for c in row):
            continue

        def _g(key, _row=row):
            idx = col.get(key)
            if idx is not None and idx < len(_row) and _row[idx] is not None:
                return str(_row[idx]).strip()
            return None

        merchant_name = _g("merchant_name")
        if not merchant_name:
            continue

        # 解析平台
        raw_platform = _g("platform")
        platform_code = None
        if raw_platform:
            platform_code = PLATFORM_NAME_MAP.get(raw_platform.strip().lower(), raw_platform.strip().upper())

        recommendations.append(MerchantRecommendation(
            mcid=_g("mcid"),
            merchant_mid=_g("merchant_mid"),
            merchant_name=merchant_name,
            platform=platform_code,
            merchant_url=_g("merchant_url"),
            merchant_region=_g("merchant_region"),
            epc=_to_float(_g("epc")),
            commission_cap=_to_float(_g("commission_cap")),
            avg_commission_rate=_to_float(_g("avg_commission_rate")),
            avg_order_commission=_to_float(_g("avg_order_commission")),
            upload_batch=batch_id,
        ))

    if not recommendations:
        raise HTTPException(status_code=400, detail="未解析到有效的推荐商家记录")

    db.add_all(recommendations)
    db.flush()

    # 标记 affiliate_merchants
    marked_count = 0
    for r in recommendations:
        conditions = []
        if r.merchant_mid and r.merchant_mid.isdigit():
            conditions.append(AffiliateMerchant.merchant_id == r.merchant_mid)
        if r.mcid:
            # 优先用 mcid 字段精确匹配
            if r.platform:
                conditions.append(
                    (func.lower(AffiliateMerchant.mcid) == r.mcid.lower())
                    & (AffiliateMerchant.platform == r.platform)
                )
            else:
                conditions.append(func.lower(AffiliateMerchant.mcid) == r.mcid.lower())
        if r.merchant_name:
            conditions.append(func.lower(AffiliateMerchant.merchant_name) == r.merchant_name.lower())

        matched = db.query(AffiliateMerchant).filter(or_(*conditions)).all()
        for m in matched:
            if m.recommendation_status != "recommended":
                m.recommendation_status = "recommended"
                m.recommendation_time = datetime.utcnow()
                marked_count += 1

    db.commit()
    return {
        "success": True,
        "batch_id": batch_id,
        "total_records": len(recommendations),
        "marked_merchants": marked_count,
    }


@router.get("")
async def list_recommendations(
    batch_id: Optional[str] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """查询推荐商家列表"""
    q = db.query(MerchantRecommendation)
    if batch_id:
        q = q.filter(MerchantRecommendation.upload_batch == batch_id)
    if search:
        like = f"%{search}%"
        q = q.filter(or_(
            MerchantRecommendation.merchant_name.ilike(like),
            MerchantRecommendation.mcid.ilike(like),
            MerchantRecommendation.merchant_mid.ilike(like),
        ))

    total = q.count()
    items = (
        q.order_by(MerchantRecommendation.id.desc())
        .offset((page - 1) * page_size).limit(page_size).all()
    )
    return {
        "total": total, "page": page, "page_size": page_size,
        "items": [{
            "id": r.id, "mcid": r.mcid, "merchant_mid": r.merchant_mid,
            "merchant_name": r.merchant_name, "merchant_url": r.merchant_url,
            "recommend_reason": r.recommend_reason,
            "merchant_region": r.merchant_region,
            "epc": float(r.epc) if r.epc else None,
            "commission_cap": float(r.commission_cap) if r.commission_cap else None,
            "avg_commission_rate": float(r.avg_commission_rate) if r.avg_commission_rate else None,
            "avg_order_commission": float(r.avg_order_commission) if r.avg_order_commission else None,
            "upload_batch": r.upload_batch,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        } for r in items],
    }


@router.get("/batches")
async def list_batches(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """查询所有上传批次"""
    batches = (
        db.query(
            MerchantRecommendation.upload_batch,
            func.count(MerchantRecommendation.id).label("count"),
            func.min(MerchantRecommendation.created_at).label("uploaded_at"),
        )
        .group_by(MerchantRecommendation.upload_batch)
        .order_by(func.min(MerchantRecommendation.created_at).desc())
        .all()
    )
    return [{
        "batch_id": b.upload_batch, "count": b.count,
        "uploaded_at": b.uploaded_at.isoformat() if b.uploaded_at else None,
    } for b in batches]
