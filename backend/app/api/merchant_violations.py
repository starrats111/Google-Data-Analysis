"""
商家违规记录 API — Excel 上传、查询、员工分配检查
"""
import uuid
import logging
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.database import get_db
from app.middleware.auth import get_current_user, get_current_manager_or_leader
from app.models.merchant import AffiliateMerchant, MerchantAssignment
from app.models.merchant_violation import MerchantViolation
from app.models.violation_report import ViolationReport
from app.models.sheet_config import SheetConfig
from app.models.notification import Notification
from app.models.user import User

router = APIRouter(prefix="/api/merchant-violations", tags=["商家违规"])
logger = logging.getLogger(__name__)

PLATFORM_NAME_MAP = {
    "linkhaitao": "LH", "lh": "LH",
    "rewardoo": "RW", "rw": "RW",
    "collabglow": "CG", "cg": "CG",
    "brandsparkhub": "BSH", "bsh": "BSH",
    "partnermatic": "PM", "pm": "PM",
    "creatorflare": "CF", "cf": "CF",
    "linkbux": "LB", "lb": "LB",
}


def _normalize_platform(raw: str) -> str:
    if not raw:
        return raw
    return PLATFORM_NAME_MAP.get(raw.strip().lower(), raw.strip().upper())


@router.post("/upload")
async def upload_violations(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_manager_or_leader),
    db: Session = Depends(get_db),
):
    """上传违规商家 Excel，解析入库，标记违规商家，检查员工分配并发通知"""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")

    import openpyxl
    from io import BytesIO

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="文件大小不能超过 10MB")

    try:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法读取 Excel: {e}")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 无数据行")

    headers = [str(h or "").strip() for h in rows[0]]
    col = {}
    for i, h in enumerate(headers):
        hl = h.lower()
        if "mcid" in hl:
            col["mcid"] = i
        elif "violation" in hl and "time" in hl:
            col["violation_time"] = i
        elif "id" in hl and ("商家" in h or "merchant" in hl):
            col["merchant_mid"] = i
        elif "名称" in h or ("name" in hl and "merchant" in hl):
            col["merchant_name"] = i
        elif "平台" in h or "platform" in hl:
            col["platform"] = i
        elif "url" in hl:
            col["merchant_url"] = i

    if "merchant_name" not in col:
        raise HTTPException(status_code=400, detail="Excel 缺少商家名称列")

    batch_id = f"VIO-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6]}"
    violations = []

    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue

        def _g(key, _row=row):
            idx = col.get(key)
            if idx is not None and idx < len(_row) and _row[idx] is not None:
                return str(_row[idx]).strip()
            return None

        merchant_name = _g("merchant_name")
        if not merchant_name:
            continue

        platform_raw = _g("platform")
        vt_raw = _g("violation_time")
        violation_time = None
        if vt_raw:
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S"):
                try:
                    violation_time = datetime.strptime(vt_raw, fmt)
                    break
                except ValueError:
                    continue

        violations.append(MerchantViolation(
            mcid=_g("mcid"),
            merchant_mid=_g("merchant_mid"),
            merchant_name=merchant_name,
            platform=_normalize_platform(platform_raw) if platform_raw else "",
            merchant_url=_g("merchant_url"),
            violation_time=violation_time,
            upload_batch=batch_id,
        ))

    if not violations:
        raise HTTPException(status_code=400, detail="未解析到有效的违规记录")

    db.add_all(violations)
    db.flush()

    # 标记 affiliate_merchants + 检查员工分配
    marked_count = 0
    alert_assignments = []

    for v in violations:
        conditions = []
        if v.merchant_mid and v.merchant_mid.isdigit():
            conditions.append(
                (AffiliateMerchant.merchant_id == v.merchant_mid)
                & (AffiliateMerchant.platform == v.platform)
            )
        if v.mcid:
            # 优先用 mcid 字段精确匹配
            conditions.append(
                (func.lower(AffiliateMerchant.mcid) == v.mcid.lower())
                & (AffiliateMerchant.platform == v.platform)
            )
        if v.merchant_name:
            conditions.append(
                (func.lower(AffiliateMerchant.merchant_name) == v.merchant_name.lower())
                & (AffiliateMerchant.platform == v.platform)
            )

        matched = db.query(AffiliateMerchant).filter(or_(*conditions)).all()
        for m in matched:
            if m.violation_status != "violated":
                m.violation_status = "violated"
                m.violation_time = v.violation_time or datetime.utcnow()
                marked_count += 1

            actives = (
                db.query(MerchantAssignment)
                .filter(MerchantAssignment.merchant_id == m.id, MerchantAssignment.status == "active")
                .all()
            )
            for a in actives:
                user = db.query(User).get(a.user_id)
                if user:
                    alert_assignments.append({
                        "merchant_name": m.merchant_name,
                        "platform": m.platform,
                        "user_id": a.user_id,
                        "user_display_name": user.display_name or user.username,
                    })

    # 发通知
    notified_users = set()
    if alert_assignments:
        user_alerts = {}
        for a in alert_assignments:
            user_alerts.setdefault(a["user_id"], []).append(f"{a['merchant_name']}({a['platform']})")

        for uid, mlist in user_alerts.items():
            s = "、".join(mlist[:5])
            if len(mlist) > 5:
                s += f" 等{len(mlist)}个"
            db.add(Notification(user_id=uid, type="violation_alert",
                                title="违规商家告警",
                                content=f"你分配的以下商家已被标记为违规：{s}。请及时处理。"))
            notified_users.add(uid)

        managers = db.query(User).filter(User.role.in_(["admin", "manager"])).all()
        affected_names = list({a["user_display_name"] for a in alert_assignments})
        emp_str = "、".join(affected_names[:5])
        if len(affected_names) > 5:
            emp_str += f" 等{len(affected_names)}人"
        for mgr in managers:
            if mgr.id not in notified_users:
                db.add(Notification(
                    user_id=mgr.id, type="violation_alert",
                    title="违规商家上传告警",
                    content=f"本次上传 {len(violations)} 条违规记录，标记 {marked_count} 个商家，"
                            f"涉及员工：{emp_str}（共 {len(alert_assignments)} 条分配）。",
                ))

    db.commit()
    return {
        "success": True, "batch_id": batch_id,
        "total_records": len(violations), "marked_merchants": marked_count,
        "affected_assignments": len(alert_assignments),
        "notified_users": len(notified_users),
        "alert_details": alert_assignments[:20],
    }


@router.get("")
async def list_violations(
    platform: Optional[str] = None,
    batch_id: Optional[str] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """查询违规商家列表"""
    q = db.query(MerchantViolation)
    if platform:
        q = q.filter(MerchantViolation.platform == _normalize_platform(platform))
    if batch_id:
        q = q.filter(MerchantViolation.upload_batch == batch_id)
    if search:
        like = f"%{search}%"
        q = q.filter(or_(
            MerchantViolation.merchant_name.ilike(like),
            MerchantViolation.mcid.ilike(like),
            MerchantViolation.merchant_mid.ilike(like),
        ))

    total = q.count()
    items = (
        q.order_by(MerchantViolation.violation_time.desc().nullslast(), MerchantViolation.id.desc())
        .offset((page - 1) * page_size).limit(page_size).all()
    )
    return {
        "total": total, "page": page, "page_size": page_size,
        "items": [{
            "id": v.id, "mcid": v.mcid, "merchant_mid": v.merchant_mid,
            "merchant_name": v.merchant_name, "platform": v.platform,
            "merchant_url": v.merchant_url,
            "violation_reason": v.violation_reason,
            "violation_time": v.violation_time.isoformat() if v.violation_time else None,
            "upload_batch": v.upload_batch,
            "created_at": v.created_at.isoformat() if v.created_at else None,
        } for v in items],
    }


@router.get("/check-assignments")
async def check_violation_assignments(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """检查哪些员工分配了违规商家"""
    violated = (
        db.query(AffiliateMerchant)
        .filter(AffiliateMerchant.violation_status == "violated")
        .all()
    )
    if not violated:
        return {"total_violated": 0, "affected_assignments": []}

    results = []
    for m in violated:
        actives = (
            db.query(MerchantAssignment)
            .filter(MerchantAssignment.merchant_id == m.id, MerchantAssignment.status == "active")
            .all()
        )
        for a in actives:
            user = db.query(User).get(a.user_id)
            results.append({
                "merchant_name": m.merchant_name, "merchant_id": m.merchant_id,
                "platform": m.platform,
                "violation_time": m.violation_time.isoformat() if m.violation_time else None,
                "user_id": a.user_id,
                "user_display_name": user.display_name if user else None,
                "assigned_at": a.assigned_at.isoformat() if a.assigned_at else None,
            })
    return {"total_violated": len(violated), "affected_assignments": results}


@router.get("/batches")
async def list_batches(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """查询所有上传批次"""
    batches = (
        db.query(
            MerchantViolation.upload_batch,
            func.count(MerchantViolation.id).label("count"),
            func.min(MerchantViolation.created_at).label("uploaded_at"),
        )
        .group_by(MerchantViolation.upload_batch)
        .order_by(func.min(MerchantViolation.created_at).desc())
        .all()
    )
    return [{
        "batch_id": b.upload_batch, "count": b.count,
        "uploaded_at": b.uploaded_at.isoformat() if b.uploaded_at else None,
    } for b in batches]


# ============================================================
# 共享表格同步
# ============================================================

class SheetUrlRequest(BaseModel):
    sheet_url: str
    config_type: str = "violation"  # violation / recommendation


@router.post("/sheet-config")
async def save_sheet_config(
    req: SheetUrlRequest,
    current_user: User = Depends(get_current_manager_or_leader),
    db: Session = Depends(get_db),
):
    """保存共享表格链接"""
    from app.services.sheet_sync_service import extract_sheet_id
    if not extract_sheet_id(req.sheet_url):
        raise HTTPException(status_code=400, detail="无效的 Google Sheets 链接")

    cfg = db.query(SheetConfig).filter(SheetConfig.config_type == req.config_type).first()
    if cfg:
        cfg.sheet_url = req.sheet_url
        cfg.updated_by = current_user.id
    else:
        cfg = SheetConfig(config_type=req.config_type, sheet_url=req.sheet_url, updated_by=current_user.id)
        db.add(cfg)
    db.commit()
    return {"success": True, "config_type": req.config_type}


@router.get("/sheet-config")
async def get_sheet_config(
    config_type: str = Query("violation"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取共享表格链接配置"""
    cfg = db.query(SheetConfig).filter(SheetConfig.config_type == config_type).first()
    if not cfg:
        return {"sheet_url": "", "last_synced_at": None}
    return {
        "sheet_url": cfg.sheet_url,
        "last_synced_at": cfg.last_synced_at.isoformat() if cfg.last_synced_at else None,
    }


@router.post("/sheet-sync")
async def sync_from_sheet(
    config_type: str = Query("violation"),
    current_user: User = Depends(get_current_manager_or_leader),
    db: Session = Depends(get_db),
):
    """从共享表格实时同步数据"""
    cfg = db.query(SheetConfig).filter(SheetConfig.config_type == config_type).first()
    if not cfg or not cfg.sheet_url:
        raise HTTPException(status_code=400, detail="未配置共享表格链接")

    try:
        if config_type == "violation":
            from app.services.sheet_sync_service import sync_violation_sheet
            result = sync_violation_sheet(db, cfg.sheet_url)
        else:
            from app.services.sheet_sync_service import sync_recommendation_sheet
            result = sync_recommendation_sheet(db, cfg.sheet_url)
    except Exception as e:
        logger.error("[SheetSync] 同步失败: %s", e)
        raise HTTPException(status_code=500, detail=f"同步失败: {str(e)}")

    return {"success": True, **result}


# ============================================================
# 员工违规上报
# ============================================================

class ViolationReportRequest(BaseModel):
    merchant_name: str
    mcid: Optional[str] = None
    merchant_mid: Optional[str] = None
    platform: Optional[str] = None
    merchant_url: Optional[str] = None
    reason: str


@router.post("/report")
async def submit_violation_report(
    req: ViolationReportRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """员工提交违规上报"""
    report = ViolationReport(
        reporter_id=current_user.id,
        merchant_name=req.merchant_name,
        mcid=req.mcid,
        merchant_mid=req.merchant_mid,
        platform=req.platform,
        merchant_url=req.merchant_url,
        reason=req.reason,
        status="pending",
    )
    db.add(report)
    db.commit()

    # 通知 leader/manager
    leaders = db.query(User).filter(User.role.in_(["manager", "leader"]), User.is_active == True).all()
    for leader in leaders:
        db.add(Notification(
            user_id=leader.id, type="violation_report",
            title="新违规上报待审核",
            content=f"{current_user.display_name or current_user.username} 上报了商家「{req.merchant_name}」违规，原因：{req.reason[:100]}",
        ))
    db.commit()

    return {"success": True, "report_id": report.id}


@router.get("/reports")
async def list_violation_reports(
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """查询违规上报列表"""
    q = db.query(ViolationReport)
    if status:
        q = q.filter(ViolationReport.status == status)
    # 普通员工只看自己的，leader/manager 看所有
    if current_user.role not in ("manager", "leader"):
        q = q.filter(ViolationReport.reporter_id == current_user.id)

    total = q.count()
    items = q.order_by(ViolationReport.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    result = []
    for r in items:
        reporter = db.query(User).get(r.reporter_id)
        reviewer = db.query(User).get(r.reviewer_id) if r.reviewer_id else None
        result.append({
            "id": r.id,
            "merchant_name": r.merchant_name,
            "mcid": r.mcid,
            "merchant_mid": r.merchant_mid,
            "platform": r.platform,
            "merchant_url": r.merchant_url,
            "reason": r.reason,
            "status": r.status,
            "reporter": reporter.display_name or reporter.username if reporter else "未知",
            "reporter_id": r.reporter_id,
            "reviewer": reviewer.display_name or reviewer.username if reviewer else None,
            "review_comment": r.review_comment,
            "reviewed_at": r.reviewed_at.isoformat() if r.reviewed_at else None,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    return {"items": result, "total": total}


class ReviewRequest(BaseModel):
    action: str  # approve / reject
    comment: Optional[str] = None


@router.post("/reports/{report_id}/review")
async def review_violation_report(
    report_id: int,
    req: ReviewRequest,
    current_user: User = Depends(get_current_manager_or_leader),
    db: Session = Depends(get_db),
):
    """审核违规上报"""
    report = db.query(ViolationReport).get(report_id)
    if not report:
        raise HTTPException(status_code=404, detail="上报记录不存在")
    if report.status != "pending":
        raise HTTPException(status_code=400, detail="该上报已审核")

    report.status = "approved" if req.action == "approve" else "rejected"
    report.reviewer_id = current_user.id
    report.review_comment = req.comment
    report.reviewed_at = datetime.utcnow()

    if req.action == "approve":
        # 写入违规记录
        batch_id = f"REPORT-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6]}"
        v = MerchantViolation(
            mcid=report.mcid,
            merchant_mid=report.merchant_mid,
            merchant_name=report.merchant_name,
            platform=report.platform or "",
            merchant_url=report.merchant_url,
            violation_reason=report.reason,
            violation_time=datetime.utcnow(),
            upload_batch=batch_id,
        )
        db.add(v)

        # 标记 AffiliateMerchant
        conditions = []
        if report.mcid:
            conditions.append(func.lower(AffiliateMerchant.mcid) == report.mcid.lower())
        if report.merchant_mid and report.merchant_mid.isdigit():
            conditions.append(AffiliateMerchant.merchant_id == report.merchant_mid)
        if report.merchant_name:
            conditions.append(func.lower(AffiliateMerchant.merchant_name) == report.merchant_name.lower())
        if conditions:
            matched = db.query(AffiliateMerchant).filter(or_(*conditions)).all()
            for m in matched:
                if m.violation_status != "violated":
                    m.violation_status = "violated"
                    m.violation_time = datetime.utcnow()

        # 尝试写入共享表格
        cfg = db.query(SheetConfig).filter(SheetConfig.config_type == "violation").first()
        if cfg and cfg.sheet_url:
            try:
                from app.services.sheet_sync_service import write_violation_to_sheet
                reporter = db.query(User).get(report.reporter_id)
                write_violation_to_sheet(
                    cfg.sheet_url, report.merchant_name, report.mcid or "",
                    report.platform or "", report.reason,
                    reporter.display_name or reporter.username if reporter else "未知"
                )
            except Exception as e:
                logger.warning("[ViolationReport] 写入共享表格失败: %s", e)

    # 通知上报人
    db.add(Notification(
        user_id=report.reporter_id, type="violation_review",
        title=f"违规上报{'已通过' if req.action == 'approve' else '已驳回'}",
        content=f"你上报的商家「{report.merchant_name}」违规{'已通过审核' if req.action == 'approve' else '被驳回'}。"
                + (f"审核意见：{req.comment}" if req.comment else ""),
    ))

    db.commit()
    return {"success": True, "status": report.status}
