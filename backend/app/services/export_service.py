"""
导出服务
将分析结果导出为Excel格式（.xlsx），使用表6作为模板格式
"""
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None

from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
from sqlalchemy.orm import Session
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.analysis_result import AnalysisResult
from app.models.affiliate_account import AffiliateAccount
from app.models.user import User


class ExportService:
    """导出服务类"""
    
    def __init__(self, export_folder: str = "excel", template_file: str = "excel/输出表格.xlsx"):
        self.export_folder = Path(export_folder)
        self.template_file = Path(template_file)
        self.export_folder.mkdir(parents=True, exist_ok=True)
    
    def export_analysis_results(
        self,
        db: Session,
        user: User,
        account_id: Optional[int] = None,
        platform_id: Optional[int] = None,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        group_by: Optional[str] = None
    ) -> str:
        """
        导出分析结果
        
        参数:
            db: 数据库会话
            user: 当前用户
            account_id: 联盟账号ID（可选）
            platform_id: 平台ID（可选）
            start_date: 开始日期（可选）
            end_date: 结束日期（可选）
            group_by: 分组方式（可选：'employee', 'platform', 'account'）
        
        返回:
            导出文件的路径
        """
        # 构建查询
        query = db.query(AnalysisResult)
        
        # 权限控制：员工只能看自己的数据
        if user.role == 'employee':
            query = query.filter(AnalysisResult.user_id == user.id)
        
        # 筛选条件
        if account_id:
            query = query.filter(AnalysisResult.affiliate_account_id == account_id)
        
        if platform_id:
            query = query.join(AffiliateAccount).filter(
                AffiliateAccount.platform_id == platform_id
            )
        
        if start_date:
            query = query.filter(AnalysisResult.analysis_date >= start_date)
        
        if end_date:
            query = query.filter(AnalysisResult.analysis_date <= end_date)
        
        # 获取数据
        results = query.all()
        
        if not results:
            raise ValueError("没有可导出的数据")
        
        if not PANDAS_AVAILABLE:
            raise ValueError("pandas 未安装，导出功能暂时不可用。请先安装 pandas 和 numpy。")
        
        # 转换为DataFrame
        data_list = []
        for result in results:
            # 获取账号信息
            account = db.query(AffiliateAccount).filter(
                AffiliateAccount.id == result.affiliate_account_id
            ).first()
            
            # 解析result_data（JSONB）
            result_data = result.result_data if isinstance(result.result_data, dict) else {}
            
            # 构建导出数据行
            for row in result_data.get('data', []):
                export_row = {
                    '日期': result.analysis_date.strftime('%Y-%m-%d') if result.analysis_date else '',
                    '员工': user.username if user.role == 'employee' else result.user.username,
                    '联盟平台': account.platform.platform_name if account and account.platform else '',
                    '联盟账号': account.account_name if account else '',
                    **row  # 展开分析结果数据
                }
                data_list.append(export_row)
        
        # 创建DataFrame
        df = pd.DataFrame(data_list)
        
        # 生成文件名
        filename = self._generate_filename(user, start_date, end_date, account_id, platform_id)
        filepath = self.export_folder / filename
        
        # 使用表6模板格式导出
        if self.template_file.exists():
            # 如果表6模板存在，使用模板格式
            self._export_with_template(df, filepath)
        else:
            # 否则使用默认格式
            self._export_default(df, filepath)
        
        return str(filepath)
    
    def _generate_filename(
        self,
        user: User,
        start_date: Optional[str],
        end_date: Optional[str],
        account_id: Optional[int],
        platform_id: Optional[int]
    ) -> str:
        """生成导出文件名"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if user.role == 'manager':
            prefix = '工作室_分析结果'
        else:
            prefix = f'{user.username}_分析结果'
        
        # 日期范围
        date_range = ''
        if start_date and end_date:
            date_range = f'_{start_date}至{end_date}'
        elif start_date:
            date_range = f'_{start_date}起'
        elif end_date:
            date_range = f'_至{end_date}'
        
        # 筛选条件
        filter_suffix = ''
        if account_id:
            filter_suffix = f'_账号{account_id}'
        if platform_id:
            filter_suffix = f'_平台{platform_id}'
        
        filename = f'{prefix}{date_range}{filter_suffix}_{timestamp}.xlsx'
        return filename
    
    def _export_with_template(self, df: pd.DataFrame, filepath: Path):
        """
        使用表6模板格式导出
        
        如果表6模板存在，读取模板的格式（列名、样式等），然后填充数据
        """
        try:
            # 读取表6模板
            template_wb = load_workbook(self.template_file)
            template_ws = template_wb.active
            
            # 获取模板的列名（第一行）
            template_columns = []
            for cell in template_ws[1]:
                template_columns.append(cell.value if cell.value else '')
            
            # 创建新工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "分析结果"
            
            # 复制表头样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入表头（使用模板的列名，如果df中有对应列则使用，否则留空）
            for col_idx, col_name in enumerate(template_columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # 写入数据
            for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
                for col_idx, col_name in enumerate(template_columns, 1):
                    # 尝试从df中获取对应列的数据
                    value = row_data.get(col_name, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # 设置数据格式
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                    elif isinstance(value, datetime):
                        cell.number_format = 'yyyy-mm-dd'
            
            # 调整列宽
            for col_idx in range(1, len(template_columns) + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 15
            
            # 保存文件
            wb.save(filepath)
            
        except Exception as e:
            # 如果使用模板失败，回退到默认导出
            print(f"使用模板导出失败: {e}，使用默认格式")
            self._export_default(df, filepath)
    
    def _export_default(self, df: pd.DataFrame, filepath: Path):
        """
        默认导出格式（如果表6模板不存在或读取失败）
        """
        # 使用pandas直接导出
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='分析结果')
            
            # 获取工作表
            worksheet = writer.sheets['分析结果']
            
            # 设置表头样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # 调整列宽
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def export_by_template_structure(self, df: pd.DataFrame, filepath: Path):
        """
        根据表6的实际结构导出
        
        需要先分析表6的列结构，然后按照该结构组织数据
        """
        # 读取表6模板获取列结构
        template_df = pd.read_excel(self.template_file, engine='openpyxl')
        template_columns = template_df.columns.tolist()
        
        # 创建新DataFrame，按照表6的列顺序
        export_df = pd.DataFrame()
        
        # 映射数据列到表6的列
        column_mapping = {
            # 根据实际表6的列名进行映射
            # 例如：'日期': '日期', '点击': '点击', '订单数': '订单数'等
        }
        
        # 如果df中有表6的列，则使用；否则创建空列
        for col in template_columns:
            if col in df.columns:
                export_df[col] = df[col]
            elif col in column_mapping and column_mapping[col] in df.columns:
                export_df[col] = df[column_mapping[col]]
            else:
                export_df[col] = ''
        
        # 导出
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            export_df.to_excel(writer, index=False, sheet_name='分析结果')
            
            # 应用样式（参考表6的样式）
            worksheet = writer.sheets['分析结果']
            self._apply_template_styles(worksheet, len(template_columns))
        
        return str(filepath)
    
    def _apply_template_styles(self, worksheet, num_columns: int):
        """应用表6模板的样式"""
        # 表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 应用表头样式
        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 应用数据行样式
        for row in worksheet.iter_rows(min_row=2, max_col=num_columns):
            for cell in row:
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                elif isinstance(cell.value, str) and cell.value and 'ROI' in str(cell.value):
                    # ROI列可能需要百分比格式
                    pass
        
        # 调整列宽
        for col in range(1, num_columns + 1):
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = 15


